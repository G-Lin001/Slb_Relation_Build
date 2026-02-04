import re
import textfsm  # type: ignore
import pandas as pd  # type: ignore
import os
from pathlib import Path
from openpyxl import load_workbook  # type: ignore
from openpyxl.styles import PatternFill  # type: ignore


# 使用TextFSM提取log内容
def extract_content_with_textfsm(template_file, input_file):
    results = {}

    if isinstance(template_file, list):
        for name, temp_file in template_file:
            print(f"\n=== 进行{name}处理 ===")
            with (
                open(temp_file, encoding="utf-8") as template,
                open(input_file, encoding="utf-8") as input_log,
            ):
                fsm = textfsm.TextFSM(template)
                Extract_data = fsm.ParseText(input_log.read())
                Extract_header = fsm.header
                # 针对dp SLB VS配置，补全缺失的VS_TYPE字段
                if "dp" in name:
                    # 定义需要处理的字段及其默认值
                    field_defaults = {
                        "VS_TYPE": "layer4",
                        "VS_AVAILABLE_STATUS": "disable",
                        "MEMBER_AVAILABLE_STATUS": "enable",
                    }
                    # 遍历需要处理的字段
                    for field_name, default_value in field_defaults.items():
                        if field_name in Extract_header:
                            field_index = Extract_header.index(field_name)
                            for i, row in enumerate(Extract_data):
                                if not row[field_index]:  # 如果字段值为空
                                    # Extract_data[i] = list(row)  # 如果TextFSM返回的是元组，则转换为列表以便修改
                                    Extract_data[i][field_index] = default_value
                # if "xa" in name:
                #     # 定义需要处理的字段及其默认值
                #     field_defaults = {
                #         "VS_HEALTH_CHECK_STATUS": "None",
                #     }
                #     # 遍历需要处理的字段
                #     for field_name, default_value in field_defaults.items():
                #         if field_name in Extract_header:
                #             field_index = Extract_header.index(field_name)
                #             for i, row in enumerate(Extract_data):
                #                 if not row[field_index]:  # 如果字段值为空
                #                     # Extract_data[i] = list(row)  # 如果TextFSM返回的是元组，则转换为列表以便修改
                #                     Extract_data[i][field_index] = default_value
                # print(f"提取的列头: {Extract_header}")
                # print("提取的数据:")
                # for row in Extract_data:
                #     print(row)
            results[name] = {"header": Extract_header, "data": Extract_data}

        return results
    # 传入单个模板文件的处理逻辑（如果需要）
    # elseif isinstance(template_file, str):
    #     with (
    #         open(template_file, encoding="utf-8") as template,
    #         open(input_file, encoding="utf-8") as input_log,
    #     ):
    #         fsm = textfsm.TextFSM(template)
    #         Extract_data = fsm.ParseText(input_log.read())
    #         Extract_header = fsm.header
    #         # 针对SLB VS配置，补全缺失的VS_TYPE字段
    #         if "VS_TYPE" in Extract_header:
    #             vs_type_index = Extract_header.index("VS_TYPE")
    #             for i, row in enumerate(Extract_data):
    #                 if not row[vs_type_index]:  # 如果VS_TYPE为空
    #                     Extract_data[i] = list(row)  # 转换为列表以便修改
    #                     Extract_data[i][vs_type_index] = "layer4"
    #         if "MEMBER_STATE" in Extract_header:
    #             member_state_index = Extract_header.index("MEMBER_STATE")
    #             for i, row in enumerate(Extract_data):
    #                 if not row[member_state_index]:  # 如果MEMBER_STATE为空
    #                     Extract_data[i] = list(row)  # 转换为列表以便修改
    #                     Extract_data[i][member_state_index] = "enabled"
    #         # print(f"提取的列头: {Extract_header}")
    #         # print("提取的数据:")
    #         # for row in Extract_data:
    #         #     print(row)
    #     results["default"] = {"header": Extract_header, "data": Extract_data}
    #     return results


# 构建dp SLB关联关系
def dp_slb_relation_build(dp_conf_results, dp_status_results, count_slb_num):
    """
    使用pandas关联vs_conf<->vs_status、pool_conf<->pool_status和member_conf<->member_status配置表
    """
    # 将配置数据转换为pandas DataFrame
    dp_vs_conf_df = pd.DataFrame(
        dp_conf_results["dp_vs_conf"]["data"],
        columns=dp_conf_results["dp_vs_conf"]["header"],
    ).drop_duplicates(subset=["VS_NAME"])
    # print(dp_vs_conf_df)
    dp_pool_conf_df = pd.DataFrame(
        dp_conf_results["dp_pool_conf"]["data"],
        columns=dp_conf_results["dp_pool_conf"]["header"],
    ).drop_duplicates(subset=["POOL_NAME"])
    dp_member_conf_df = pd.DataFrame(
        dp_conf_results["dp_member_conf"]["data"],
        columns=dp_conf_results["dp_member_conf"]["header"],
    ).drop_duplicates(subset=["MEMBER_NAME"])

    # 将状态数据转换为pandas DataFrame
    dp_vs_status_df = pd.DataFrame(
        dp_status_results["dp_vs_status"]["data"],
        columns=dp_status_results["dp_vs_status"]["header"],
    ).drop_duplicates(subset=["VS_NAME"])
    # print(dp_vs_status_df)
    dp_pool_status_df = pd.DataFrame(
        dp_status_results["dp_pool_status"]["data"],
        columns=dp_status_results["dp_pool_status"]["header"],
    ).drop_duplicates(subset=["POOL_NAME"])
    # print(dp_pool_status_df)
    dp_member_status_df = pd.DataFrame(
        dp_status_results["dp_member_status"]["data"],
        columns=dp_status_results["dp_member_status"]["header"],
    ).drop_duplicates(subset=["MEMBER_NAME"])
    # print(dp_pool_conf_df["POOL_NAME"].count())
    # 初步数据完整性检查
    print("\n=== 开始进行数据完整性检查 ===")
    if dp_pool_conf_df["POOL_NAME"].count() != dp_pool_status_df["POOL_NAME"].count():
        print("警告:Pool配置和状态中的POOL_NAME数量不匹配,请检查数据完整性。")
        print(f"配置中POOL_NAME数量: {dp_pool_conf_df['POOL_NAME'].count()}")
        print(f"状态中POOL_NAME数量: {dp_pool_status_df['POOL_NAME'].count()}")
        return None
    # print(dp_pool_conf_df["POOL_NAME"].count())
    if (
        dp_member_conf_df["MEMBER_NAME"].count()
        != dp_member_status_df["MEMBER_NAME"].count()
    ):
        print("警告:Member配置和状态中的MEMBER_NAME数量不匹配,请检查数据完整性。")
        print(f"配置中MEMBER_NAME数量: {dp_member_conf_df['MEMBER_NAME'].count()}")
        print(f"状态中MEMBER_NAME数量: {dp_member_status_df['MEMBER_NAME'].count()}")
        return None
    if dp_vs_conf_df["VS_NAME"].count() != dp_vs_status_df["VS_NAME"].count():
        print("警告:VS配置和状态中的VS_NAME数量不匹配,请检查数据完整性。")
        print(f"配置中VS_NAME数量: {dp_vs_conf_df['VS_NAME'].count()}")
        print(f"状态中VS_NAME数量: {dp_vs_status_df['VS_NAME'].count()}")
        return None
    print("模板提取后的配置和状态表内部数据完整性检查通过。")

    print("\n=== 开始进行日志与配置数据完整性检查 ===")
    if (
        count_slb_num.get("dp_log_vs_count") == dp_vs_conf_df["VS_NAME"].count()
        and count_slb_num.get("dp_log_pool_count")
        == dp_pool_conf_df["POOL_NAME"].count()
        and count_slb_num.get("dp_log_member_count")
        == dp_member_conf_df["MEMBER_NAME"].count()
    ):
        print("数据完整性检查通过，日志文件中的数据数量与模板提取后的数据数量一致。")
    else:
        print(
            "警告: 数据完整性检查失败，请检查日志文件中的数据数量与配置表中的数据数量是否一致。"
        )
        print(f"日志中VS_NAME数量: {count_slb_num.get('dp_log_vs_count')}")
        print(f"配置中VS_NAME数量: {dp_vs_conf_df['VS_NAME'].count()}")
        print(f"日志中POOL_NAME数量: {count_slb_num.get('dp_log_pool_count')}")
        print(f"配置中POOL_NAME数量: {dp_pool_conf_df['POOL_NAME'].count()}")
        print(f"日志中MEMBER_NAME数量: {count_slb_num.get('dp_log_member_count')}")
        print(f"配置中MEMBER_NAME数量: {dp_member_conf_df['MEMBER_NAME'].count()}")
        return None

    # print("=== 原始数据预览 ===")
    # print("VS配置:")
    # print(dp_vs_conf_df.head())
    # print("\nPool配置:")
    # print(dp_pool_conf_df.head())
    # print("\nMember配置:")
    # print(dp_member_conf_df.head())
    # print("\nVS状态:")
    # print(dp_vs_status_df.head())
    # print("\nPool状态:")
    # print(dp_pool_status_df.head())
    # print("\nMember状态:")
    # print(dp_member_status_df.head())

    # 检查关联列是否存在
    vs_name_col = (
        "VS_NAME"
        if "VS_NAME" in dp_vs_conf_df.columns and "VS_NAME" in dp_vs_status_df.columns
        else None
    )
    # print(f"VS_NAME_COL: {vs_name_col}")
    pool_name_col = (
        "POOL_NAME"
        if "POOL_NAME" in dp_pool_conf_df.columns
        and "POOL_NAME" in dp_pool_status_df.columns
        and "POOL_NAME" in dp_vs_conf_df.columns
        else None
    )
    # print(f"POOL_NAME_COL: {pool_name_col}")
    member_name_col = (
        "MEMBER_NAME"
        if "MEMBER_NAME" in dp_member_conf_df.columns
        and "MEMBER_NAME" in dp_member_status_df.columns
        and "MEMBER_NAME" in dp_pool_conf_df.columns
        else None
    )
    # print(f"MEMBER_NAME_COL: {member_name_col}")

    if vs_name_col is None or pool_name_col is None or member_name_col is None:
        print("错误：存在未找到必要的关联列，请检查列头信息。")
        print(f"vs_conf列: {list(dp_vs_conf_df.columns)}")
        print(f"pool_conf列: {list(dp_pool_conf_df.columns)}")
        print(f"member_conf列: {list(dp_member_conf_df.columns)}")
        print(f"vs_status列: {list(dp_vs_status_df.columns)}")
        print(f"pool_status列: {list(dp_pool_status_df.columns)}")
        print(f"member_status列: {list(dp_member_status_df.columns)}")
        return None
    else:
        print("所有必要的关联列均已找到，开始进行数据关联处理。")

    # 处理vs_conf中的POOL_NAME列表
    # 将列表形式的POOL_NAME拆分为单个池
    print("\n=== 处理vs_conf中的POOL_NAME列表 ===")
    vs_expanded_list = []

    for _, vs_row in dp_vs_conf_df.iterrows():
        pool_names = vs_row["POOL_NAME"]
        for pool_name in pool_names:
            # 创建新的行，复制vs的所有信息，但POOL_NAME改为单个池
            new_row = vs_row.copy()
            new_row["POOL_NAME"] = pool_name
            vs_expanded_list.append(new_row)
    # 创建展开后的vs DataFrame
    dp_vs_conf_expanded_df = pd.DataFrame(vs_expanded_list)
    print("\n=== vs_conf中的POOL_NAME列表处理完成 ===")
    # print(f"展开后的VS配置行数: {len(dp_vs_conf_expanded_df)}")
    # print("展开后的VS配置:")
    # print(dp_vs_conf_expanded_df.head())

    # 处理pool_conf中的MEMBER_NAME列表
    # 将列表形式的MEMBER_NAME拆分为单个成员
    print("\n=== 处理pool_conf中的MEMBER_NAME列表 ===")
    dp_pool_expanded_list = []

    for _, pool_row in dp_pool_conf_df.iterrows():
        member_names = pool_row["MEMBER_NAME"]
        for member_name in member_names:
            # 创建新的行，复制pool的所有信息，但MEMBER_NAME改为单个成员
            new_row = pool_row.copy()
            new_row["MEMBER_NAME"] = member_name
            dp_pool_expanded_list.append(new_row)

    # 创建展开后的pool DataFrame
    pool_conf_expanded_df = pd.DataFrame(dp_pool_expanded_list)
    # print(f"展开后的Pool配置行数: {len(pool_conf_expanded_df)}")
    # print("展开后的Pool配置:")
    # print(pool_conf_expanded_df.head())
    # pool_expanded_df.to_excel(
    #     "d:/Content Extraction/dp_out_files/expanded_pool_conf.xlsx", index=False
    # )
    print("\n=== pool_conf中的MEMBER_NAME列表处理完成 ===")

    # 链式合并DataFrame
    print("\n=== 开始进行数据关联处理 ===")
    final_merged = (
        dp_vs_conf_expanded_df.merge(
            dp_vs_status_df,
            left_on="VS_NAME",
            right_on="VS_NAME",
            how="left",
            suffixes=("", "_vs_status"),
        )
        .merge(
            pool_conf_expanded_df,
            left_on="POOL_NAME",
            right_on="POOL_NAME",
            how="left",
            suffixes=("", "_pool"),
        )
        .merge(
            dp_pool_status_df,
            left_on="POOL_NAME",
            right_on="POOL_NAME",
            how="left",
            suffixes=("", "_pool_status"),
        )
        .merge(
            dp_member_conf_df,
            left_on="MEMBER_NAME",
            right_on="MEMBER_NAME",
            how="left",
            suffixes=("", "_member"),
        )
        .merge(
            dp_member_status_df,
            left_on="MEMBER_NAME",
            right_on="MEMBER_NAME",
            how="left",
            suffixes=("", "_member_status"),
        )
    )
    print("=== 数据关联处理完成 ===")

    print("\n=== dp slb conf配置处理后统计信息 ===")
    print(f"VS数量: {dp_vs_conf_df['VS_NAME'].nunique()}")
    print(f"Pool数量: {dp_pool_conf_df['POOL_NAME'].nunique()}")
    print(f"Member数量: {dp_member_conf_df['MEMBER_NAME'].nunique()}")

    print("\n=== dp slb status状态处理后统计信息 ===")
    print(f"VS数量: {dp_vs_status_df['VS_NAME'].nunique()}")
    print(f"Pool数量: {dp_pool_status_df['POOL_NAME'].nunique()}")
    print(f"Member数量: {dp_member_status_df['MEMBER_NAME'].nunique()}")

    print("\n=== dp slb关联完成后配置统计信息 ===")
    print(f"VS数量: {final_merged['VS_NAME'].nunique()}")
    print(f"Pool数量: {final_merged['POOL_NAME'].nunique()}")
    print(f"Member数量: {final_merged['MEMBER_NAME'].nunique()}")

    print("\n=== dp slb未使用配置统计信息 ===")
    print(
        f"\n未使用POOL数量:{int(dp_pool_conf_df['POOL_NAME'].nunique()) - int(final_merged['POOL_NAME'].nunique())}"
    )
    print(
        f"未使用Member数量: {int(dp_member_conf_df['MEMBER_NAME'].nunique()) - int(final_merged['MEMBER_NAME'].nunique())}"
    )
    return final_merged


# 构建xa SLB关联关系
def xa_slb_relation_build(xa_results, count_slb_num):
    """
    处理xa_results中的数据,构建关联关系
    :return: None
    """
    # xa的show tech 文件存在重复数据，处理时需要对df关键字段去重
    # 将xa_results转换为DataFrame
    xa_vs_status_df = pd.DataFrame(
        xa_results["xa_vs_status"]["data"],
        columns=xa_results["xa_vs_status"]["header"],
    ).drop_duplicates(subset=["VS_NAME"], keep="first")
    # 处理xa_vs_ssl
    xa_vs_ssl_df = pd.DataFrame(
        xa_results["xa_vs_ssl"]["data"],
        columns=xa_results["xa_vs_ssl"]["header"],
    ).drop_duplicates(subset=["VS_NAME", "SSL_NAME"], keep="first")
    # 处理xa_vs_pool_relation
    xa_vs_pool_relation_df = pd.DataFrame(
        xa_results["xa_vs_pool_relation"]["data"],
        columns=xa_results["xa_vs_pool_relation"]["header"],
    ).drop_duplicates(subset=["VS_NAME", "POOL_NAME"], keep="first")
    # 处理xa_vs_member_relation
    xa_vs_member_relation_df = pd.DataFrame(
        xa_results["xa_vs_member_relation"]["data"],
        columns=xa_results["xa_vs_member_relation"]["header"],
    ).drop_duplicates(subset=["VS_NAME", "MEMBER_NAME"], keep="first")
    # 处理xa_pool_member_relation
    xa_pool_member_relation_df = pd.DataFrame(
        xa_results["xa_pool_member_relation"]["data"],
        columns=xa_results["xa_pool_member_relation"]["header"],
    ).drop_duplicates(subset=["POOL_NAME", "MEMBER_NAME"], keep="first")
    # 处理xa_member_status
    xa_member_status_df = pd.DataFrame(
        xa_results["xa_member_status"]["data"],
        columns=xa_results["xa_member_status"]["header"],
    ).drop_duplicates(subset=["MEMBER_NAME"], keep="first")

    print("\n=== 开始进行数据完整性检查 ===")
    if (
        count_slb_num.get("xa_log_vs_count") == xa_vs_status_df["VS_NAME"].count()
        and count_slb_num.get("xa_log_pool_use_count")
        == xa_pool_member_relation_df["POOL_NAME"].nunique()
        and count_slb_num.get("xa_log_member_count")
        == xa_member_status_df["MEMBER_NAME"].count()
        # and count_slb_num.get("xa_log_ssl_count") == xa_vs_ssl_df["SSL_NAME"].nunique()
    ):
        print("数据完整性检查通过，日志文件中的数据数量与模板提取后的数据数量一致。")
    else:
        print(
            "警告: 数据完整性检查失败，请检查日志文件中的数据数量与配置表中的数据数量是否一致。"
        )
        print(f"日志中VS_NAME数量: {count_slb_num.get('xa_log_vs_count')}")
        print(f"模板中VS_NAME数量: {xa_vs_status_df['VS_NAME'].count()}")
        print(f"日志中POOL_NAME数量: {count_slb_num.get('xa_log_pool_use_count')}")
        print(
            f"模板中POOL_NAME数量: {xa_pool_member_relation_df['POOL_NAME'].nunique()}"
        )
        print(f"日志中MEMBER_NAME数量: {count_slb_num.get('xa_log_member_count')}")
        print(f"模板中MEMBER_NAME数量: {xa_member_status_df['MEMBER_NAME'].count()}")
        print(f"日志中SSL_NAME数量: {count_slb_num.get('xa_log_ssl_count')}")
        print(f"模板中SSL_NAME数量: {xa_vs_ssl_df['SSL_NAME'].nunique()}")
        return None

    # 表链接逻辑处理
    # 1. 存在VS-POOL这层级关系，通过VS关联Pool，再通过Pool关联Member
    # 2. 丢弃没关联到POOL的VS数据
    print("\n=== 开始进行数据关联处理 ===")
    xa_vs_pool_member_merged = (
        xa_vs_status_df.merge(
            xa_vs_pool_relation_df,
            left_on="VS_NAME",
            right_on="VS_NAME",
            how="left",
            suffixes=("", "_vs_pool"),
        )
        .merge(
            xa_pool_member_relation_df,
            left_on="POOL_NAME",
            right_on="POOL_NAME",
            how="left",
            suffixes=("", "_pool_member"),
        )
        .merge(
            xa_member_status_df,
            left_on="MEMBER_NAME",
            right_on="MEMBER_NAME",
            how="left",
            suffixes=("", "_member_status"),
        )
    ).dropna(subset=["POOL_NAME"])

    # 1. 将全量VS和VS-MEMBER关系表关联
    # 2. 丢弃没关联到MEMBER的VS数据
    xa_vs_member_merged = (
        xa_vs_status_df.merge(
            xa_vs_member_relation_df,
            left_on="VS_NAME",
            right_on="VS_NAME",
            how="left",
            suffixes=("", "_vs_member"),
        ).merge(
            xa_member_status_df,
            left_on="MEMBER_NAME",
            right_on="MEMBER_NAME",
            how="left",
            suffixes=("", "_member_status"),
        )
    ).dropna(subset=["MEMBER_NAME"])

    # 合并两部分数据,并且关联上SSL_HOST
    final_merged = pd.concat([xa_vs_pool_member_merged, xa_vs_member_merged]).merge(
        xa_vs_ssl_df,
        left_on="VS_NAME",
        right_on="VS_NAME",
        how="left",
        suffixes=("", "_vs_ssl"),
    )

    print("=== 数据关联处理完成 ===")

    print("\n=== xa slb配置提取后配置统计信息 ===")
    print(f"VS数量: {xa_vs_status_df['VS_NAME'].nunique()}")
    print(f"Pool数量: {xa_pool_member_relation_df['POOL_NAME'].nunique()}")
    print(f"Member数量: {xa_member_status_df['MEMBER_NAME'].nunique()}")

    print("\n=== xa slb关联完成后配置统计信息 ===")
    print(f"VS数量: {final_merged['VS_NAME'].nunique()}")
    print(f"Pool数量: {final_merged['POOL_NAME'].nunique()}")
    print(f"Member数量: {final_merged['MEMBER_NAME'].nunique()}")

    print("\n=== xa slb未使用配置统计信息 ===")
    print(
        f"POOL数量:{int(xa_pool_member_relation_df['POOL_NAME'].nunique()) - int(final_merged['POOL_NAME'].nunique())}"
    )
    print(
        f"Member数量: {int(xa_member_status_df['MEMBER_NAME'].nunique()) - int(final_merged['MEMBER_NAME'].nunique())}"
    )

    return final_merged


# 构建hj SLB关联关系
def hj_slb_relation_build(hj_results, count_slb_num):
    hj_vs_pool_status_df = pd.DataFrame(
        hj_results["hj_vs_pool_status"]["data"],
        columns=hj_results["hj_vs_pool_status"]["header"],
    ).drop_duplicates(subset=["VS_NAME"], keep="first")

    hj_pool_member_status_df = pd.DataFrame(
        hj_results["hj_pool_member_status"]["data"],
        columns=hj_results["hj_pool_member_status"]["header"],
    ).drop_duplicates(subset=["POOL_NAME", "MEMBER_IP", "MEMBER_PORT"], keep="first")

    print("\n=== 弘积负载不进行数据完整性检查,请注意核对数量 ===")
    print("日志与模板提取数据对比信息如下:")
    print(f"日志中VS_NAME数量: {count_slb_num.get('hj_log_vs_count')}")
    print(f"模板中VS_NAME数量: {hj_vs_pool_status_df['VS_NAME'].count()}")
    print(f"日志中POOL_NAME数量: {count_slb_num.get('hj_log_pool_count')}")
    print(f"模板中POOL_NAME数量: {hj_pool_member_status_df['POOL_NAME'].nunique()}")
    print(f"日志中MEMBER_NAME数量: {count_slb_num.get('hj_log_member_count')}")
    print(f"模板中MEMBER_NAME数量: {hj_pool_member_status_df['MEMBER_IP'].count()}")

    print("\n=== 开始进行数据关联处理 ===")
    final_merged = hj_vs_pool_status_df.merge(
        hj_pool_member_status_df,
        left_on="POOL_NAME",
        right_on="POOL_NAME",
        how="left",
        suffixes=("", "_pool_member"),
    )
    return final_merged


# 统计slb配置数量
def count_slb_num_in_log(log_file_path, log_type):
    """
    根据日志类型自动选择相应模式并统计匹配次数

    Args:
        log_file_path: 日志文件路径
        log_type: 日志类型 ("dp" 或 "xa")

    Returns:
        dict: 模式名称及其命中次数的字典
    """
    # 根据日志类型选择相应的模式配置
    if log_type.lower() == "dp":
        patterns_config = [
            (
                "dp_log_vs_count",
                r"^slb virtual-service ([^ ]+)",
            ),
            (
                "dp_log_pool_count",
                r"^slb pool ([^ ]+)",
            ),
            (
                "dp_log_member_count",
                r"^slb member ([^ ]+)",
            ),
        ]
    elif log_type.lower() == "xa":
        patterns_config = [
            (
                "xa_log_vs_count",
                r"^(?:tcp|tcps|udp|http|https|dns|dnstcp|ftp|ftps|rdp|siptcp|sip|udp|radauth|radacct|diameter) virtual service.*",
            ),
            (
                "xa_log_ssl_count",
                r"^ssl host (?:virtual|real) ([^ ]+)",
            ),
            (
                "xa_log_pool_use_count",
                r"^slb group member ([^ ]+)",
            ),
            (
                "xa_log_member_count",
                r"^Real service.*",
            ),
        ]
    elif log_type.lower() == "hj":
        patterns_config = [
            (
                "hj_log_vs_count",  # 匹配virtual-address 配置中的vs name的数量
                r"^      name ([^ ]+)",
            ),
            (
                "hj_log_pool_count",  # 匹配virtual-address 状态中被关联的pool数量
                r"^-+ POOL: ([^ ]+)",
            ),
            (
                "hj_log_member_count",
                r"(^\S+:\S+)(?=\s+curr)",
            ),
        ]
    else:
        print(f"未知的日志类型: {log_type}")
        return {}

    # 读取日志文件内容
    try:
        with open(log_file_path, "r", encoding="utf-8") as file:
            log_content = file.read()
    except FileNotFoundError:
        print(f"错误: 找不到文件 {log_file_path}")
        return {}
    except Exception as e:
        print(f"读取文件时发生错误: {e}")
        return {}

    results = {}

    # 对每个模式进行匹配并统计
    for name, pattern in patterns_config:
        try:
            matches = re.findall(pattern, log_content, re.MULTILINE | re.IGNORECASE)
            unique_matches = list(set(matches))
            # print(unique_matches)
            results[name] = len(unique_matches)
        except re.error as e:
            print(f"正则表达式错误 ({pattern}): {e}")
            results[name] = 0

    # 输出结果
    return results


# 辅助函数：处理slb数据关联
def process_slb_data(log_pairs, log_type, templates, status_templates=None):
    """
    通用的SLB数据处理函数

    Args:
        log_pairs: 日志文件对列表
        log_type: 日志类型 ('dp', 'xa', 'hj')
        templates: 模板文件列表
        status_templates: dp类型的状态模板列表（可选）
    """
    if log_pairs:
        print(f"\n*************** 开始进行{log_type}数据关联 ***************")

        for log_pair in log_pairs:
            device_name = log_pair.get("device_name")

            if log_type == "dp":
                # DP类型特殊处理，需要两个日志文件
                status_log = log_pair["status_log"]
                conf_log = log_pair["conf_log"]

                print(f"\n*************** 开始处理设备: {device_name} ***************")
                count_slb_num = count_slb_num_in_log(status_log, log_type)

                # 处理DP状态日志和配置日志
                dp_status_results = extract_content_with_textfsm(
                    status_templates, status_log
                )
                dp_conf_results = extract_content_with_textfsm(templates, conf_log)

                # 构建DP SLB关联
                slb_relation_df = dp_slb_relation_build(
                    dp_conf_results, dp_status_results, count_slb_num
                )
            else:
                # XA和HJ类型处理
                log_file = log_pair[f"{log_type}_log"]

                print(f"\n*************** 开始处理设备: {device_name} ***************")
                count_slb_num = count_slb_num_in_log(log_file, log_type)

                # 处理日志文件
                results = extract_content_with_textfsm(templates, log_file)

                # 根据类型选择对应的关联构建函数
                if log_type == "xa":
                    slb_relation_df = xa_slb_relation_build(results, count_slb_num)
                elif log_type == "hj":
                    slb_relation_df = hj_slb_relation_build(results, count_slb_num)

            # 导出结果
            export_result(
                slb_relation_df,
                log_type,
                f"{log_type}_slb关联_{device_name}",
                device_name,
            )
            print(f"*************** 设备: {device_name} 处理完成 ***************\n")

        print(f"*************** {log_type}数据关联处理完成 ***************\n")
    else:
        print(f"没有提供{log_type}日志文件,跳过{log_type}数据关联处理。")


# 辅助函数：从目录获取日志文件对
def get_log_pairs_from_directory(directory_path, log_type):
    """
    从指定目录自动读取所有日志文件并生成log_pairs

    Args:
        directory_path: 日志文件所在目录路径
        log_type: 日志类型 ('dp', 'xa', 'hj')

    Returns:
        list: 生成的log_pairs列表
    """
    if not os.path.exists(directory_path):
        print(f"错误：目录 {directory_path} 不存在")
        return []

    log_pairs = []

    if log_type.lower() == "dp":
        # 对于dp类型，我们需要配对conf和status文件
        conf_files = []
        status_files = []

        # 首先收集所有文件
        for filename in os.listdir(directory_path):
            file_path = os.path.join(directory_path, filename)
            if os.path.isfile(file_path):
                if filename.endswith("-conf"):
                    conf_files.append(file_path)
                elif filename.endswith("-slb"):
                    status_files.append(file_path)

        # 然后配对文件
        for conf_file in conf_files:
            # 提取设备名（去掉-conf后缀）
            device_name = os.path.basename(conf_file).replace("-conf", "")

            # 寻找对应的status文件
            status_file = None
            for sf in status_files:
                if device_name in sf:
                    status_file = sf
                    break

            if status_file:
                log_pairs.append(
                    {
                        "status_log": status_file,
                        "conf_log": conf_file,
                        "device_name": device_name,
                    }
                )
            else:
                print(f"警告：找不到与 {conf_file} 对应的status文件")

    elif log_type.lower() == "xa" or log_type.lower() == "hj":
        # 对于xa和hj类型，直接读取所有文件
        for filename in os.listdir(directory_path):
            file_path = os.path.join(directory_path, filename)
            if os.path.isfile(file_path):
                # 提取设备名
                device_name = os.path.basename(file_path).split(".")[0]

                if log_type.lower() == "xa":
                    log_pairs.append({"xa_log": file_path, "device_name": device_name})
                else:
                    log_pairs.append({"hj_log": file_path, "device_name": device_name})

    return log_pairs


# 辅助函数：导出结果到Excel
def export_result(df, prefix, process_name, device_name=None):
    """辅助函数：导出结果到Excel"""
    if df is None:
        print(f"{process_name}过程中出现错误,程序终止。")
    else:
        print(f"\n=== {process_name}完成,开始导出结果到Excel ===")

        # 如果提供了设备名，则在输出文件名中包含设备名
        if device_name:
            output_filename = f"linked_slb_{device_name}.xlsx"
            print(f"结果已导出,文件名称: {output_filename}")
        else:
            output_filename = "linked_slb.xlsx"
            print("结果已导出,文件名称: linked_slb.xlsx")

        df.to_excel(
            f"d:/Content Extraction/Out_Files/{prefix}_Out_Files/{output_filename}",
            index=False,
        )


# 通用的Excel文件合并函数
def merge_excel_files(
    log_type, input_source, output_path, file_pattern="*.xlsx", sheet_name="Merged_Data"
):
    """
    合并Excel文件，可以处理目录路径或文件路径列表

    Args:
        input_source: 目录路径或文件路径列表
        output_path: 输出文件路径
        file_pattern: 文件匹配模式（仅当input_source为目录时有效）
        sheet_name: 工作表名称
    """
    print("=== 开始合并Excel文件 ===")

    # 处理输入源：如果是目录，获取所有Excel文件；如果是列表，直接使用

    # 从目录获取文件
    directory = Path(input_source)
    excel_files = list(directory.glob(file_pattern))
    excel_files.extend(list(directory.glob("*.xls")))  # 也包括.xls文件

    if not excel_files:
        print(f"在目录 {input_source} 中未找到Excel文件")
        return False
    else:
        file_paths = [str(file) for file in excel_files]
        print(f"从目录 {input_source} 找到 {len(file_paths)} 个Excel文件")

    # 打印要合并的文件名
    for file_path in file_paths:
        print(f"  - {os.path.basename(file_path)}")

    # 记录每个文件的行数和数据
    file_info = []
    all_dataframes = []

    # 读取所有文件
    for i, file_path in enumerate(file_paths, 1):
        print(f"正在处理第 {i}/{len(file_paths)} 个文件: {os.path.basename(file_path)}")
        try:
            df = pd.read_excel(file_path)
            rows_count = len(df)

            # 添加源文件标识列
            device_name = (
                os.path.basename(file_path)
                .replace("linked_slb_", "")
                .replace(".xlsx", "")
            )
            df.insert(0, "Device_Name", device_name)

            file_info.append({"file": os.path.basename(file_path), "rows": rows_count})
            all_dataframes.append(df)
            print(f"  - 读取行数: {rows_count}")
        except Exception as e:
            print(f"  - 读取文件失败: {e}")
            continue

    # 检查是否有数据可以合并
    if not all_dataframes:
        print("没有有效的数据可以合并")
        return False

    # 计算总行数
    total_rows_original = sum(info["rows"] for info in file_info)
    print(f"\n原始总行数: {total_rows_original}")

    # 合并所有DataFrame
    merged_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
    total_rows_merged = len(merged_df)
    print(f"合并后总行数: {total_rows_merged}")

    # 验证行数
    if total_rows_original == total_rows_merged:
        print("✓ 数据行数验证通过！")
    else:
        print(
            f"⚠ 警告: 数据行数不匹配！原始: {total_rows_original}, 合并后: {total_rows_merged}"
        )

    # 保存合并后的文件
    print(f"\n正在保存到: {output_path}")
    merged_df.to_excel(output_path, sheet_name=sheet_name, index=False)

    # 显示结果
    print(f"\n=== {log_type}类型Excel文件合并完成！ ===")
    print(f"总共合并了 {len(file_info)} 个文件")
    print(f"输出文件: {output_path}")

    # 显示详细信息
    print("\n各文件详情:")
    for info in file_info:
        print(f"  - {info['file']}: {info['rows']} 行")

    return True


# 合并指定类型的SLB Excel文件
def merge_slb_excel_files(log_type, output_filename):
    """
    合并指定类型的SLB Excel文件

    Args:
        log_type: SLB类型，支持"dp", "xa", "hj"
        output_filename: 输出文件名（不含路径）
    """
    # 构建输入目录路径
    input_directory = f"D:\\Content Extraction\\Out_Files\\{log_type}_Out_Files"

    # 检查目录是否存在
    if not os.path.exists(input_directory):
        print(f"{log_type}类型的输出目录不存在: {input_directory}")
        return False

    # 构建输出路径
    output_path = f"D:\\Content Extraction\\Out_Files\\{output_filename}"
    if not output_path.endswith((".xlsx", ".xls")):
        output_path += ".xlsx"

    # 执行合并
    print(f"\n=== 开始合并{log_type}类型的Excel文件 ===")
    success = merge_excel_files(log_type, input_directory, output_path)

    if success:
        print(f"=== {log_type}类型的Excel文件合并完成！输出文件: {output_path} ===")

    return output_path


# dp SLB状态分析处理
def dp_slb_status_analyze(
    input_file,
    output_file,
    group_column="VS_NAME",
    available_status_column="MEMBER_AVAILABLE_STATUS",
    health_status_column="MEMBER_HEALTH_CHECK_STATUS",
):
    """
    根据VS_NAME列分组，并分析每组中MEMBER_AVAILABLE_STATUS和MEMBER_HEALTH_CHECK_STATUS列的数据，
    然后根据逻辑判断添加状态列，并对特定列进行颜色格式化。

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出文件名
        group_column: 分组依据的列名，默认为'VS_NAME'
        available_status_column: 可用性状态列名，默认为'MEMBER_AVAILABLE_STATUS'
        health_status_column: 健康检查状态列名，默认为'MEMBER_HEALTH_CHECK_STATUS'
    """
    # 1. 读取Excel文件
    df = pd.read_excel(input_file)

    # 复制数据框以避免修改原始数据
    df_modified = df.copy()

    # 添加新的状态列，初始化为空字符串
    df_modified["处理建议"] = ""

    # 按照group_column进行分组
    grouped = df_modified.groupby(group_column)
    print("开始分析dp SLB状态...")
    # 遍历每个组
    for group_name, group_df in grouped:
        # 计算当前组的行数
        total_rows = len(group_df)

        # 统计MEMBER_AVAILABLE_STATUS列中'enable'的数量
        enable_count = (group_df[available_status_column] == "enable").sum()

        # 统计MEMBER_HEALTH_CHECK_STATUS列中'Up'的数量
        up_count = (group_df[health_status_column] == "Up").sum()

        # 根据逻辑判断设置状态
        if enable_count == total_rows and up_count == total_rows:
            status_text = "正常"
        elif 0 < enable_count < total_rows and up_count == total_rows:
            status_text = "部分mem未启用，请确认是否保留"
        elif enable_count == total_rows and 0 < up_count < total_rows:
            status_text = "部分mem探测down，请确认是否保留"
        elif 0 < enable_count < total_rows and 0 < up_count < total_rows:
            status_text = "部分mem探测down或未启用，请确认是否保留"
        elif enable_count == 0 or up_count == 0:
            status_text = "所有mem探测down或未启用，请确认是否保留"
        else:
            status_text = "未知状态"

        # 更新当前组所有行的Tips列
        df_modified.loc[group_df.index, "处理建议"] = status_text

    # 将结果保存到Excel文件
    df_modified.to_excel(output_file, index=False)

    # 加载工作簿以应用颜色格式
    wb = load_workbook(output_file)
    ws = wb.active

    # 获取列索引
    available_status_col_idx = df_modified.columns.get_loc(available_status_column) + 1
    health_status_col_idx = df_modified.columns.get_loc(health_status_column) + 1

    # 定义颜色填充
    green_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 遍历所有数据行（跳过标题行）
    for row in range(2, len(df_modified) + 2):
        # 对MEMBER_AVAILABLE_STATUS列进行颜色格式化
        available_cell = ws.cell(row=row, column=available_status_col_idx)
        if available_cell.value == "enable":
            available_cell.fill = green_fill
        else:
            available_cell.fill = red_fill

        # 对MEMBER_HEALTH_CHECK_STATUS列进行颜色格式化
        health_cell = ws.cell(row=row, column=health_status_col_idx)
        if health_cell.value == "Up":
            health_cell.fill = green_fill
        else:
            health_cell.fill = red_fill

    # 保存工作簿
    wb.save(output_file)
    print("分析完成！结果已保存到:", output_file)
    return df_modified


# xa SLB状态分析处理
def xa_slb_status_analyze(
    input_file,
    output_file,
    group_column="VS_NAME",
    available_status_column="MEMBER_AVAILABLE_STATUS",
    health_status_column="MEMBER_HEALTH_CHECK_STATUS",
):
    """
    根据VS_NAME列分组，并分析每组中MEMBER_AVAILABLE_STATUS和MEMBER_HEALTH_CHECK_STATUS列的数据，
    然后根据逻辑判断添加状态列，并对特定列进行颜色格式化。

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出文件名
        group_column: 分组依据的列名，默认为'VS_NAME'
        available_status_column: 可用性状态列名，默认为'MEMBER_AVAILABLE_STATUS'
        health_status_column: 健康检查状态列名，默认为'MEMBER_HEALTH_CHECK_STATUS'
    """
    # 1. 读取Excel文件
    df = pd.read_excel(input_file)

    # 复制数据框以避免修改原始数据
    df_modified = df.copy()

    # 添加新的状态列，初始化为空字符串
    df_modified["处理建议"] = ""

    # 按照group_column进行分组
    grouped = df_modified.groupby(group_column)
    print("开始分析xa SLB状态...")
    # 遍历每个组
    for group_name, group_df in grouped:
        # 计算当前组的行数
        total_rows = len(group_df)

        # 统计MEMBER_AVAILABLE_STATUS列中'ACTIVE'的数量
        active_count = (group_df[available_status_column] == "ACTIVE").sum()

        # 统计MEMBER_HEALTH_CHECK_STATUS列中'UP'的数量
        up_count = (group_df[health_status_column] == "UP").sum()

        # 根据逻辑判断设置状态
        if active_count == total_rows and up_count == total_rows:
            status_text = "正常"
        elif 0 < active_count < total_rows and up_count == total_rows:
            status_text = "部分mem未启用，请确认是否保留"
        elif active_count == total_rows and 0 < up_count < total_rows:
            status_text = "部分mem探测down，请确认是否保留"
        elif 0 < active_count < total_rows and 0 < up_count < total_rows:
            status_text = "部分mem探测down或未启用，请确认是否保留"
        elif active_count == 0 or up_count == 0:
            status_text = "所有mem探测down或未启用，请确认是否保留"
        else:
            status_text = "未知状态"

        # 更新当前组所有行的Tips列
        df_modified.loc[group_df.index, "处理建议"] = status_text

    # 将结果保存到Excel文件
    df_modified.to_excel(output_file, index=False)

    # 加载工作簿以应用颜色格式
    wb = load_workbook(output_file)
    ws = wb.active

    # 获取列索引
    available_status_col_idx = df_modified.columns.get_loc(available_status_column) + 1
    health_status_col_idx = df_modified.columns.get_loc(health_status_column) + 1

    # 定义颜色填充
    green_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 遍历所有数据行（跳过标题行）
    for row in range(2, len(df_modified) + 2):
        # 对MEMBER_AVAILABLE_STATUS列进行颜色格式化
        available_cell = ws.cell(row=row, column=available_status_col_idx)
        if available_cell.value == "ACTIVE":
            available_cell.fill = green_fill
        else:
            available_cell.fill = red_fill

        # 对MEMBER_HEALTH_CHECK_STATUS列进行颜色格式化
        health_cell = ws.cell(row=row, column=health_status_col_idx)
        if health_cell.value == "UP":
            health_cell.fill = green_fill
        else:
            health_cell.fill = red_fill

    # 保存工作簿
    wb.save(output_file)
    print("分析完成！结果已保存到:", output_file)
    return df_modified


# hj SLB状态分析处理
def hj_slb_status_analyze(
    slb_status_in_file,
    slb_status_out_file,
    vs_column="VS_NAME",
    mem_health_status_column="MEMBER_HEALTH_CHECK_STATUS",
):
    """
    根据VS_NAME列分组，分析MEMBER_HEALTH_CHECK_STATUS列的数据，
    添加状态列并对健康检查状态列进行颜色格式化。

    Args:
        input_file: 输入Excel文件路径
        output_file: 输出Excel文件路径
        vs_column: VS_NAME列名，默认为'VS_NAME'
        health_status_column: 健康检查状态列名，默认为'MEMBER_HEALTH_CHECK_STATUS'
    """
    # 1. 读取Excel文件
    df = pd.read_excel(slb_status_in_file)

    # 复制数据框以避免修改原始数据
    df_modified = df.copy()

    # 添加新的状态列，初始化为空字符串
    df_modified["处理建议"] = ""

    # 2. 按照VS_NAME进行分组
    grouped = df_modified.groupby(vs_column)
    print("开始分析hj SLB状态...")
    # 3. 遍历每个组进行分析
    for group_name, group_df in grouped:
        # 计算当前组的行数
        total_rows = len(group_df)

        # 提取MEMBER_HEALTH_CHECK_STATUS列中UP的数量
        up_count = (group_df[mem_health_status_column] == "UP").sum()

        # 4. 进行逻辑判断并设置状态
        if up_count == total_rows:
            status_text = "正常"
        elif 0 < up_count < total_rows:
            status_text = "部分mem探测down或未启用，请确认是否保留"
        elif up_count == 0:
            status_text = "所有mem探测down或未启用，请确认是否保留"
        else:
            status_text = "未知状态"

        # 更新当前组所有行的状态列
        df_modified.loc[group_df.index, "处理建议"] = status_text

    # 将结果保存到Excel文件
    df_modified.to_excel(slb_status_out_file, index=False)

    # 加载工作簿以应用颜色格式
    wb = load_workbook(slb_status_out_file)
    ws = wb.active

    # 获取健康检查状态列的索引
    health_status_col_idx = df_modified.columns.get_loc(mem_health_status_column) + 1

    # 定义颜色填充
    green_fill = PatternFill(
        start_color="92D050", end_color="92D050", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # 5. 对MEMBER_HEALTH_CHECK_STATUS列进行颜色格式化
    for row in range(2, len(df_modified) + 2):
        health_cell = ws.cell(row=row, column=health_status_col_idx)
        if health_cell.value == "UP":
            health_cell.fill = green_fill
        else:
            health_cell.fill = red_fill

    # 保存工作簿
    wb.save(slb_status_out_file)

    print("分析完成！结果已保存到:", slb_status_out_file)
    return df_modified


# slb status分析处理函数，用于简化调用
def process_slb_status_analyze(
    log_type,
    slb_status_in_file,
    slb_status_out_file,
    group_column="VS_NAME",
    mem_available_status_column="MEMBER_AVAILABLE_STATUS",
    mem_health_status_column="MEMBER_HEALTH_CHECK_STATUS",
):
    """
    统一的SLB建议处理函数，根据数据类型自动选择合适的处理方法。

    Args:
        log_type: 日志类型，"dp", "xa" 或 "hj"
        input_file: 输入文件名
        output_file: 输出文件名
        group_column: 分组依据的列名，默认为'VS_NAME'
        available_status_column: 可用性状态列名，默认为'MEMBER_AVAILABLE_STATUS'
        health_status_column: 健康检查状态列名，默认为'MEMBER_HEALTH_CHECK_STATUS'

    Returns:
        处理后的DataFrame
    """
    if log_type == "dp":
        # 调用dp的处理函数
        return dp_slb_status_analyze(
            slb_status_in_file,
            slb_status_out_file,
            group_column,
            mem_available_status_column,
            mem_health_status_column,
        )
    elif log_type == "xa":
        # 调用dp的处理函数
        return xa_slb_status_analyze(
            slb_status_in_file,
            slb_status_out_file,
            group_column,
            mem_available_status_column,
            mem_health_status_column,
        )
    elif log_type == "hj":
        # 调用hj的处理函数
        return hj_slb_status_analyze(
            slb_status_in_file,
            slb_status_out_file,
            group_column,
            mem_health_status_column,
        )
    else:
        raise ValueError("data_type must be either 'dp', 'xa' or 'hj'")


def main():
    dp_conf_templates = [
        (
            "dp_vs_conf",
            "d:/Content Extraction/Templates/dp_Templates/vs/dp_slb_vs_conf.textfsm",
        ),
        (
            "dp_pool_conf",
            "d:/Content Extraction/Templates/dp_Templates/pool/dp_slb_pool_conf.textfsm",
        ),
        (
            "dp_member_conf",
            "d:/Content Extraction/Templates/dp_Templates/member/dp_slb_member_conf.textfsm",
        ),
    ]

    dp_status_templates = [
        (
            "dp_vs_status",
            "d:/Content Extraction/Templates/dp_Templates/vs/dp_slb_vs_status.textfsm",
        ),
        (
            "dp_pool_status",
            "d:/Content Extraction/Templates/dp_Templates/pool/dp_slb_pool_status.textfsm",
        ),
        (
            "dp_member_status",
            "d:/Content Extraction/Templates/dp_Templates/member/dp_slb_member_status.textfsm",
        ),
    ]

    xa_templates = [
        (
            "xa_vs_status",
            "d:/Content Extraction/Templates/xa_Templates/vs/xa_slb_vs_status.textfsm",
        ),
        (
            "xa_vs_pool_relation",
            "d:/Content Extraction/Templates/xa_Templates/vs/xa_slb_vs_pool_relation.textfsm",
        ),
        (
            "xa_vs_member_relation",
            "d:/Content Extraction/Templates/xa_Templates/vs/xa_slb_vs_member_relation.textfsm",
        ),
        (
            "xa_vs_ssl",
            "d:/Content Extraction/Templates/xa_Templates/vs/xa_slb_vs_ssl.textfsm",
        ),
        (
            "xa_pool_member_relation",
            "d:/Content Extraction/Templates/xa_Templates/pool/xa_slb_pool_member_relation.textfsm",
        ),
        (
            "xa_member_status",
            "d:/Content Extraction/Templates/xa_Templates/member/xa_slb_member_status.textfsm",
        ),
    ]

    hj_templates = [
        (
            "hj_vs_pool_status",
            "d:/Content Extraction/Templates/hj_Templates/vs/hj_slb_vs_pool_status.textfsm",
        ),
        (
            "hj_pool_member_status",
            "d:/Content Extraction/Templates/hj_Templates/pool/hj_slb_pool_member_status.textfsm",
        ),
    ]

    # 自动从目录读取dp日志文件
    dp_directory = "d:/Content Extraction/Log/dp_Log"
    dp_log_pairs = get_log_pairs_from_directory(dp_directory, "dp")

    # 自动从目录读取xa日志文件
    xa_directory = "d:/Content Extraction/Log/xa_Log"
    xa_log_pairs = get_log_pairs_from_directory(xa_directory, "xa")

    # 自动从目录读取hj日志文件
    hj_directory = "d:/Content Extraction/Log/hj_Log"
    hj_log_pairs = get_log_pairs_from_directory(hj_directory, "hj")

    # # 处理dp_slb数据关联
    # process_slb_data(dp_log_pairs, "dp", dp_conf_templates, dp_status_templates)

    # # 处理xa_slb数据关联
    process_slb_data(xa_log_pairs, "xa", xa_templates)

    # 处理hj_slb数据关联
    process_slb_data(hj_log_pairs, "hj", hj_templates)

    # # 合并三种类型的Excel文件
    # merge_slb_excel_files("dp", "dp_slb_健康检查状态表.xlsx")
    xa_slb_status_file = merge_slb_excel_files("xa", "xa_slb_健康检查状态表.xlsx")
    hj_slb_status_file = merge_slb_excel_files("hj", "hj_slb_健康检查状态表.xlsx")

    # 进行三种类型的SLB状态分析
    process_slb_status_analyze(
        "xa",
        xa_slb_status_file,
        "d:/Content Extraction/Out_Files/xa_slb_健康检查状态表_分析后.xlsx",
    )
    process_slb_status_analyze(
        "hj",
        hj_slb_status_file,
        "d:/Content Extraction/Out_Files/hj_slb_健康检查状态表_分析后.xlsx",
    )


if __name__ == "__main__":
    main()
